"""Helper functions: dates, overdue logic, file handling, Excel export, etc."""

from __future__ import annotations

import io
import os
import re
import shutil
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from utils.constants import (
    STATUS_COLORS,
    DEFAULT_COLOR,
    TERMINAL_STATUSES,
    PLATFORMS,
)
from utils.i18n import t

# Import Supabase client (centralized in db layer)
try:
    from db.database import get_supabase
except Exception:
    get_supabase = None  # will fail gracefully if called without proper setup

# ------------------------------------------------------------------
# Date & Overdue
# ------------------------------------------------------------------
def parse_date(d: Optional[str]) -> Optional[date]:
    if not d:
        return None
    try:
        return datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def format_date(d: Optional[str], default: str = "-") -> str:
    """Format ISO date string for display (YYYY-MM-DD remains clean)."""
    dt = parse_date(d)
    if dt is None:
        return default
    return dt.isoformat()


def is_overdue(deadline: Optional[str], status: str) -> bool:
    """Return True if deadline has passed and status is not terminal."""
    if status in TERMINAL_STATUSES:
        return False
    dt = parse_date(deadline)
    if dt is None:
        return False
    return dt < date.today()


def get_overdue_flag(deadline: Optional[str], status: str, lang: str = "zh") -> str:
    return t("overdue_flag", lang) if is_overdue(deadline, status) else t("not_overdue", lang)


# ------------------------------------------------------------------
# Status helpers
# ------------------------------------------------------------------
def get_status_color(status: str) -> str:
    return STATUS_COLORS.get(status, DEFAULT_COLOR)


def get_platforms() -> list[str]:
    return PLATFORMS.copy()


# ------------------------------------------------------------------
# File / Attachment helpers - now using Supabase Storage
# ------------------------------------------------------------------
def safe_filename(name: str) -> str:
    """Sanitize filename, keep extension."""
    name = name.strip().replace(" ", "_")
    name = re.sub(r"[^\w\-\.\u4e00-\u9fff]", "", name)  # keep CJK + alnum + - .
    if not name:
        name = "file"
    # Limit length
    if len(name) > 120:
        base, ext = os.path.splitext(name)
        name = base[:110] + ext
    return name


def upload_file_to_storage(uploaded_file, supplier_id: int) -> Tuple[str, str]:
    """
    Upload Streamlit UploadedFile to Supabase Storage bucket "uploads".
    Path format: "{supplier_id}/{timestamp}_{safe_name}"
    Returns (original_filename, stored_storage_path)
    """
    if get_supabase is None:
        raise RuntimeError("Supabase client not available")

    sb = get_supabase()
    original = uploaded_file.name
    safe = safe_filename(original)
    timestamp = int(time.time() * 1000)
    stem, ext = os.path.splitext(safe)
    stored_name = f"{timestamp}_{stem}{ext}"
    storage_path = f"{supplier_id}/{stored_name}"

    # Upload bytes
    file_bytes = uploaded_file.getbuffer()
    content_type = "application/octet-stream"
    if ext.lower() in [".pdf"]:
        content_type = "application/pdf"
    elif ext.lower() in [".png", ".jpg", ".jpeg"]:
        content_type = "image/jpeg" if ext.lower() in [".jpg", ".jpeg"] else "image/png"

    sb.storage.from_("uploads").upload(
        path=storage_path,
        file=file_bytes,
        file_options={"content-type": content_type, "upsert": "true"},
    )

    return original, storage_path


# Backwards-compatible aliases for existing call sites
save_uploaded_file = upload_file_to_storage


def get_file_bytes(stored_path: str) -> bytes:
    """Download file bytes from Supabase Storage (replaces local get_full_path + open)."""
    if get_supabase is None:
        raise RuntimeError("Supabase client not available")
    sb = get_supabase()
    try:
        return sb.storage.from_("uploads").download(stored_path)
    except Exception as e:
        raise FileNotFoundError(f"Could not download {stored_path} from Supabase Storage: {e}")


def get_full_path(stored_path: str) -> Path:
    """Deprecated for Supabase. Kept for compatibility but will raise."""
    raise NotImplementedError(
        "Local file paths are no longer used. Use get_file_bytes(stored_path) instead "
        "for downloads, or access storage directly."
    )


def delete_file_from_storage(stored_path: str) -> bool:
    """Delete a single file from Supabase Storage."""
    if get_supabase is None:
        return False
    sb = get_supabase()
    try:
        sb.storage.from_("uploads").remove([stored_path])
        return True
    except Exception:
        return False


def delete_uploaded_file(stored_path: str) -> bool:
    """Alias for storage delete (back-compat with UI code)."""
    return delete_file_from_storage(stored_path)


def delete_all_files_for_supplier(supplier_id: int) -> None:
    """Delete all files for a supplier from Supabase Storage (lists under prefix)."""
    if get_supabase is None:
        return
    sb = get_supabase()
    try:
        prefix = f"{supplier_id}/"
        files = sb.storage.from_("uploads").list(prefix)
        paths = [f"{prefix}{f['name']}" for f in files if f.get("name")]
        if paths:
            sb.storage.from_("uploads").remove(paths)
    except Exception as e:
        # non-fatal
        print(f"Warning: failed to delete all storage files for supplier {supplier_id}: {e}")


def delete_all_attachments_for_supplier(supplier_id: int) -> None:
    """Back-compat alias."""
    delete_all_files_for_supplier(supplier_id)


# ------------------------------------------------------------------
# Excel Export (beautifully formatted)
# ------------------------------------------------------------------
def export_suppliers_to_excel(
    df: pd.DataFrame, lang: str = "zh", filename_hint: str = "供应商注册追踪"
) -> bytes:
    """
    Create a professionally formatted .xlsx from the suppliers dataframe.
    Returns bytes ready for st.download_button.
    """
    if df is None or df.empty:
        # Create minimal workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"
        ws["A1"] = t("no_suppliers", lang)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Prepare display dataframe (select + rename columns for export)
    export_df = df.copy()

    # Select and order nice columns (v2 includes country + owner)
    cols_order = [
        "id", "company_name_cn", "company_name_en", "country", "platform", "status",
        "submission_date", "deadline", "contact_name", "owner", "contact_email",
        "contact_phone", "notes", "created_at", "updated_at"
    ]
    existing = [c for c in cols_order if c in export_df.columns]
    export_df = export_df[existing]

    # Rename for human readability (bilingual friendly headers)
    rename_map = {
        "id": t("id", lang),
        "company_name_cn": t("company_cn", lang),
        "company_name_en": t("company_en", lang),
        "country": t("country", lang),
        "platform": t("platform", lang),
        "status": t("status", lang),
        "submission_date": t("submission_date", lang),
        "deadline": t("deadline", lang),
        "contact_name": t("contact", lang),
        "owner": t("owner", lang),
        "contact_email": "Email",
        "contact_phone": "Phone",
        "notes": t("notes", lang),
        "created_at": "Created",
        "updated_at": "Updated",
    }
    export_df = export_df.rename(columns={k: v for k, v in rename_map.items() if k in export_df.columns})

    # Add overdue flag column (computed)
    if "deadline" in df.columns and "status" in df.columns:
        overdue_series = df.apply(
            lambda r: "YES" if is_overdue(r.get("deadline"), r.get("status", "")) else "",
            axis=1,
        )
        export_df.insert(6, "Overdue", overdue_series)  # after deadline-ish position

    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Title row
    title = f"{t('export_filename', lang)} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_df.columns))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (row 3)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for col_idx, col_name in enumerate(export_df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[3].height = 22

    # Data rows
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(export_df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.font = data_font
            cell.border = thin_border

            # Status column coloring (heuristic: column named "status" or "状态")
            col_name_lower = str(export_df.columns[col_idx - 1]).lower()
            if "status" in col_name_lower or "状态" in str(export_df.columns[col_idx - 1]):
                color = STATUS_COLORS.get(str(value), DEFAULT_COLOR)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                cell.alignment = center_align
            elif col_idx == 1:  # ID
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto column widths (with sensible caps)
    for col_idx, col_name in enumerate(export_df.columns, 1):
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) if v is not None else 0 for v in export_df.iloc[:, col_idx - 1])
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)

    # Freeze header
    ws.freeze_panes = "A4"

    # Auto filter
    ws.auto_filter.ref = f"A3:{get_column_letter(len(export_df.columns))}{3 + len(export_df)}"

    # Output
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def get_export_filename(lang: str = "zh") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    base = t("export_filename", lang)
    return f"{base}_{ts}.xlsx"
